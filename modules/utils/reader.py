from openpyxl import load_workbook
from xlrd import open_workbook_xls



ABREVIATIONS = [chr(x) for x in range(ord('A'), ord('Z') + 1)]

for char in ABREVIATIONS.copy():
    for x in range(ord('A'), ord('Z') + 1):
        ABREVIATIONS.append(char + chr(x))


# print(ABREVIATIONS.index("AA"))

class XlsCell:
    """
    Since xlrd doesn't have many features like openpyxl so i decide to crate custom cell object
    
    """

    def __init__(self, coordinate, value):
        """
        Coordinate of original cell position not coodinate of ranges
        """

        self.coordinate = coordinate
        self.value = value



    def __repr__(self):
        return f"<XlsCustomCellObject at {hex(id(self))}>"



class XlsSupport:
    def __init__(self, fname):
        self._wb = open_workbook_xls(fname)
        self._ws = self._wb[0]
        self.xls_type = True

    def get_value_xls(self, coord, reverse=False):
        if reverse:
            row, col = self.convert_coord_xls(coord, reverse=True)

            return self._ws.cell(row, col).value

        row, col = self.convert_coord_xls(coord)

        return self._ws.cell(row, col).value


    def get_row(self, col, start_row, skip_empty=False):
        if skip_empty:
            return [value for value in self._ws.col_values(col, start_rowx=start_row) if value != ""]

        res = []

        cells = self._ws.col_slice(col, start_rowx=start_row)
        
        for cell in cells:
            # print(cell.value, type(cell.value),)
            if type(cell.value) is float:
                res.append(
                XlsCell(start_row, str(int(cell.value)))
                )

            else:
                res.append(
                XlsCell(self.convert_coord_xls_reverse(((start_row, col))), str(cell.value))
                )

            start_row += 1

        return res

        # return self._ws.col_slice(col, start_rowx=start_row)


    def get_col_range(self, coord_range):

        """
        Range must be like "A9:B9"

        return value of cell's from given range
        """


        row, col_start, _, col_end = self.convert_coord_xls(coord_range)


        if col_start > col_end:
            raise ValueError("start column must be less then end column!")


        return [cell.value for cell in self._ws.row_slice(row, start_colx=col_start, end_colx=col_end)]


    @classmethod
    def split_coord_ranges(cls, coord_ranges, change_row="", change_col=""):

        start_, end_ = coord_ranges.split(":")

        return [cls.split_coord(start_, change_row=change_row, change_col=change_col), cls.split_coord(end_, change_row=change_row, change_col=change_col)]


    @classmethod
    def join_coord(cls, coord, reverse=False):
        if reverse:
            coord = reversed(coord)

        return "".join(coord)


    @classmethod
    def split_coord(cls, coord, change_row="", change_col=""):
        """
        return e.g "B12" => row="12" cols="B"
        """
        cols = []
        rows = []


        for char in coord:
            if char.isdigit():
                if change_row and type(change_row) is str:
                    rows.append(change_row)
                else:
                    rows.append(char)
            else:
                if change_col and type(change_col) is str:
                    cols.append(change_col)
                else:
                    cols.append(char)

        return "".join(rows), "".join(cols)     


    def get_cell_from_ranges(self, coord_ranges):
        """
        coord_ranges must be itterable
        """

        res = []


        for coord in coord_ranges:

            start_cell, end_cell = self.split_coord_ranges(coord)
            cells = []

            # for char in range(ord(start_cell[1]), ord(end_cell[1]) + 1):
            for index in range(ABREVIATIONS.index(start_cell[1]), ABREVIATIONS.index(end_cell[1]) + 1):

                cur_coordinate = "".join((ABREVIATIONS[index], start_cell[0]))


                
                cells.append(
                    XlsCell(cur_coordinate, self.get_value_xls(cur_coordinate, reverse=True))
                )

            res.append(cells)

        return res

            


    @classmethod
    def convert_coord_ranges_xls(cls, ranges):
        """
        :ranges = splited coordinate [('8', 'E'), ('8', 'M')] usally return value from split_coord_ranges function

        return (row, col) where row and col are in integer type
        """

        cells = []



        for coord_range in ranges:
            
            cells.append(
                "".join(reversed(coord_range))
            )
        
        return ":".join(cells)



    @classmethod
    def convert_coord_xls_reverse(cls, *cells):
        """
        :cell_ => itterable object that contains (row, column)
        """

        res = []

        for cell in cells:
            res.append(
                "".join(
                    (ABREVIATIONS[cell[1]], str(cell[0]))
                )
            )

        return res


    @classmethod
    def convert_coord_xls(cls, coord, reverse=False):
        
        if len(coord.split(":")) > 1:
            start_cell, end_cell = coord.split(":")

            start_row, start_col = cls.split_coord(start_cell)
            end_row, end_col = cls.split_coord(end_cell)

            # print(end_col, ABREVIATIONS.index(end_col))

            return int(start_row) - 1, ABREVIATIONS.index(start_col), int(end_row) - 1, ABREVIATIONS.index(end_col)



        col, row = cls.split_coord(coord)


        if reverse:
            return int(col) - 1, ABREVIATIONS.index(row)

        return int(row) - 1, ABREVIATIONS.index(col)


"""
NOTE:
for year only need to specify start cell and last cell 

for eg (A2:B2)

where A2 is the start cell and the B2 is the end cell
"""


class Reader(XlsSupport):
    def __init__(self, fname, config):
        self.fname = fname
        self.config = config
        self.xls_type = False
        self._years = {}
        self._names = []
        self._codes = []
        self.loaded = False

        if not fname:
            return

        if fname.endswith(".xls"):
            super().__init__(self.fname)

        else:
            self._wb = load_workbook(self.fname)
            self._ws = self._wb.active

        self.loaded = True

    @property
    def get_copies(self):
        return self._ws.get_rows()


    @classmethod
    def wrap_year(cls, value):
        value = value.split()
        
        if len(value) != 3:
            return

        return value[-1]

    @classmethod
    def pack_months(cls, value):
        res = []

        for cell in value:
            res.append(cell[0])

        return res


    def get_months(self, update=False):
        if not self._years:
            self.get_years()

        months = []


        if self.xls_type:
            for coord in self.config["year_coords"]:        

                months.append(self.get_cell_from_ranges([self.convert_coord_ranges_xls(self.split_coord_ranges(coord, change_row="8"))])[0])

            i = 0

            for key in self._years.keys():
                self._years[key] = months[i]

                i += 1

            del months, i

            return self._years

        for coord in self.config["year_coords"]:
            row, col_start, _, col_end = self.convert_coord_xls(coord)

            months.append(tuple(self._ws.iter_cols(min_row=row+2, max_row=row+2, min_col=col_start+1, max_col=col_end+1)))


        counter = 0

        for year in self._years.copy():
            self._years[year] = self.pack_months(months[counter])

            counter += 1

        del months, counter
        return self._years


    def get_years(self, update=False):
        if self._years and not update:
            return self._years

        if self.xls_type:

            for coord in self.config["year_coords"]:
                _year = self.wrap_year(self.get_col_range(coord)[0])
                
                self._years[_year] = []

            return self._years


        for coord in self.config["year_coords"]:
            row, col_start, _, col_end = self.convert_coord_xls(coord)

            _year = tuple(self._ws.iter_cols(min_col=col_start+1, max_col=col_end+1, min_row=row+1, max_row=row+1))[0][0].value

            self._years[self.wrap_year(_year)] = []


        return self._years

    
    def get_names(self, update=False):
        if self._names and not update:
            return self._names

        if self.xls_type:
            row, col = self.convert_coord_xls(self.config["name_coords"], reverse=True)

            self._names = self.get_row(col, row)

            return self._names



        row, col = self.convert_coord_xls(self.config["name_coords"], reverse=True)

        self._names = self.pack_months(self._ws.iter_rows(min_col=col+1, max_col=col+1, min_row=row+1))

        return self._names


    def get_codes(self, update=False):
        if self._codes and not update:
            return self._codes


        if self.xls_type:
            row, col = self.convert_coord_xls(self.config["code_coords"], reverse=True)

            self._codes = self.get_row(col, row)

            return self._codes

        row, col = self.convert_coord_xls(self.config["code_coords"], reverse=True)

        self._codes = self.pack_months(self._ws.iter_rows(min_row=row+1, min_col=col+1, max_col=col+1))
        
        # print(self._codes)

        return self._codes